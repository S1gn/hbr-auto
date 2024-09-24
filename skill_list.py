# 2024-09-16
# 将作业轴行动表转化为本脚本配置的技能顺序
# 读取文件模板参考 actions.xlsx
# 行动轴工作表：turns
# 人物技能对应工作表：skill_name
import openpyxl
import argparse
from if_cmd import translate_if_cmd

def swap(lst, a, b):
    tmp = lst[a]
    lst[a] = lst[b]
    lst[b] = tmp
    return lst

def parse_opt():
    parser = argparse.ArgumentParser()
    parser.add_argument('--path', type=str, default='./actions.xlsx', help='行动轴表格路径')
    parser.add_argument('--log', type=bool, default=False, help='是否显示出日志')
    parser.add_argument('--output', type=str, default='./cmd_list.txt', help='输出文件路径')
    opt = parser.parse_args()
    return opt

def read_excel(path):
    wb = openpyxl.load_workbook(path)
    turns = wb['turns']
    skill_name1 = wb['skill_name1']
    skill_name2 = wb['skill_name2']
    return turns, skill_name1, skill_name2

def read_skill_name(skill_name):
    # B2-G2
    character_name = []
    for i in range(2, 8):
        character_name.append(str(skill_name.cell(2, i).value))
    skill_list = []
    for i in range(0, 6):
        one_skill = []
        for j in range(1, 9):
            one_skill.append(str(skill_name.cell(j + 2, i + 2).value))
        skill_list.append(one_skill)
    return character_name, skill_list

def read_turns_list(turns, turns_num, now_turns_num):
    turns_list = []
    for i in range(1, turns_num + 1):
        one_turn = []
        for j in range(1, 9):
            one_turn.append(turns.cell(i + 3 + now_turns_num, j + 2).value)
        turns_list.append(one_turn)
    return turns_list

def get_cmd(turns_list, skill_list, character_name, now_pos, if_cmd_list=None, log=False):
    cmd_list = []
    # 维护一个动态的列表，表示人物当前位置
    for i in range(0, len(turns_list)):
        if turns_list[i][6] == 1 or turns_list[i][6] == '1':
            val = 'O'
            cmd_list.append([[val]])
        count = 0
        this_cmd = [[1], [2], [3]]
        if(log):
            print(f"第{i}回合", end=' ')
            print("本轮开始位置顺序：", end=' ')
            for l in range(0, 6):
                print(character_name[now_pos[l]], end=' ')
            print('\n')
        for j in range(0, 6):
            if turns_list[i][j] == None:
                continue
            # 先遍历每个行动，有无1、2、3开头
            # 有1、2、3开头的，指令重排
            if turns_list[i][j].startswith('1') or turns_list[i][j].startswith('2') or turns_list[i][j].startswith('3'):
                # 重排，一定按照1、2、3的顺序处理，否则人物位置会出错
                ordered_cmd_list = [[], [], []]
                # 保存人物标号
                ordered_char_list = [-1, -1, -1]
                for k in range(0, 6):
                    if turns_list[i][k] == None:
                        continue
                    if turns_list[i][k].startswith('1'):
                        ordered_cmd_list[0] = turns_list[i][k]
                        ordered_char_list[0] = k
                    elif turns_list[i][k].startswith('2'):
                        ordered_cmd_list[1] = turns_list[i][k]
                        ordered_char_list[1] = k
                    elif turns_list[i][k].startswith('3'):
                        ordered_cmd_list[2] = turns_list[i][k]
                        ordered_char_list[2] = k
                # 重排后依次处理
                for k in range(0, 3):
                    char_pos = now_pos.index(ordered_char_list[k]) + 1
                    if(log):
                        print("    本条指令：", character_name[ordered_char_list[k]], "释放", turns_list[i][ordered_char_list[k]], end=' ')
                        print("    交换位置：", character_name[now_pos[k]], character_name[now_pos[char_pos - 1]])
                    
                    
                    skill_cmd = ordered_cmd_list[k].split('-')
                    if skill_cmd[1] == 'IF':
                        if_cmd = if_cmd_list[int(skill_cmd[2]) - 1]
                        trans_cmd = translate_if_cmd(if_cmd[0], if_cmd[1], if_cmd[2], skill_list, now_pos, character_name)
                        trans_cmd[0] = k + 1
                        this_cmd[k] = trans_cmd.copy()
                    else:
                        now_pos = swap(now_pos, char_pos - 1, k)
                        skill_turn = skill_list[ordered_char_list[k]].index(skill_cmd[1])
                        skill_target = 0
                        if len(skill_cmd) == 3:
                            skill_target_name = skill_cmd[2]
                            skill_target = now_pos.index(character_name.index(skill_target_name)) + 1
                        this_cmd[k] = [k + 1, char_pos, skill_turn, skill_target]
                
                # 重排后，直接跳过
                break

            # 无所谓顺序的
            else:
                # 保持位置不变，实现很麻烦，要指令重排，筛掉位置不变的，再设置顺序
                # 一律重排
                char_pos = now_pos.index(j) + 1
                # 只有不在前排的才交换
                # if  now_pos[j] > 2:
                if(log):
                        print("    本条指令：", character_name[j], "释放", turns_list[i][j], end=' ')
                        print("    交换位置：", character_name[now_pos[count]], character_name[now_pos[char_pos - 1]])
                now_pos = swap(now_pos, count, char_pos - 1)

                skill_cmd = turns_list[i][j].split('-')
                skill_turn = skill_list[j].index(skill_cmd[0])
                skill_target = 0
                if len(skill_cmd) == 2:
                    skill_target_name = skill_cmd[1]
                    skill_target = now_pos.index(character_name.index(skill_target_name)) + 1
                this_cmd[count] = [count + 1, char_pos, skill_turn, skill_target]
                count += 1
        
        # 回合中释放OD
        if turns_list[i][7] == 1 or turns_list[i][7] == '1':
            val = 'O'
            this_cmd.append([val])
            

        cmd_list.append(this_cmd.copy())
    return cmd_list, now_pos

def output_cmd(cmd_list, output):
    # 在后面接着写
    with open(output, 'a') as f:
        f.write("--------------------------------------\n")
        for i in range(0, len(cmd_list)):
            f.write("[");
            for j in range(len(cmd_list[i])):
                f.write(f"{cmd_list[i][j]}")
                if j != len(cmd_list[i]) - 1:
                    f.write(", ")
            if i != len(cmd_list) - 1:
                f.write("],\n")
            else:
                f.write("]\n")
        f.write("--------------------------------------\n")

def read_if_cmd(turns_sheet):
    # 读取if指令 i=3, j=10
    if_cmd_list = []
    i = 4
    if_cmd_text = turns_sheet.cell(i, 11).value
    if_cmd = []
    # [函数名，[目标1，目标2....]，指令]
    while if_cmd_text != None:
        cmd_name = if_cmd_text.split('(')[0]
        cmd_target = if_cmd_text.split('(')[1].split(')')[0].split(',')
        cmd_target = [x for x in cmd_target]
        cmd = if_cmd_text.split('-')[1]
        if_cmd = [cmd_name, cmd_target, cmd]
        if_cmd_list.append(if_cmd)
        i += 1
        if_cmd_text = turns_sheet.cell(i, 11).value
    return if_cmd_list

if __name__ == "__main__":
    opt = parse_opt()
    turns, skill_name1, skill_name2 = read_excel(opt.path)
    if_cmd_list = read_if_cmd(turns)

    # 清空cmd_list.txt
    with open(opt.output, 'w') as f:
        f.write("")
    # 读取最多5个回合数
    truns_num = []
    for i in range(2, 7):
        if turns.cell(1, i).value == None:
            break
        truns_num.append(int(turns.cell(1, i).value))

    now_turns_num = 0
    battle1_pos = [0, 1, 2, 3, 4, 5]
    battle2_pos = [0, 1, 2, 3, 4, 5]
    for i in range(0, len(truns_num)):
        if i % 2 == 0:
            skill_name = skill_name1
            character_name, skill_list = read_skill_name(skill_name)
            turns_list = read_turns_list(turns, truns_num[i], now_turns_num)
            cmd_list, battle1_pos = get_cmd(turns_list, skill_list, character_name, battle1_pos, if_cmd_list, opt.log)
        else:
            skill_name = skill_name2
            character_name, skill_list = read_skill_name(skill_name)
            turns_list = read_turns_list(turns, truns_num[i], now_turns_num)
            cmd_list, battle2_pos = get_cmd(turns_list, skill_list, character_name, battle2_pos, if_cmd_list, opt.log)

        output_cmd(cmd_list, "./cmd_list.txt")

        if(opt.log):
            print("指令列表：")
            for j in range(0, len(cmd_list)):
                print(cmd_list[j])

        now_turns_num += truns_num[i]